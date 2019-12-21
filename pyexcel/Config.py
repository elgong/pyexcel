
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font

# cell 背景 颜色参数字典
"""
fill_type="solid"       # 纯色填充方式
start_color='FF0000FF'  # 前景色
end_color='FF000000'    # 背景色
"""
cell_background = {
    None: PatternFill(fill_type="none", start_color='FFFFFF', end_color='FF000000'),

    "red": PatternFill(fill_type="solid", start_color='FF4500', end_color='FF000000'),

    "green":  PatternFill(fill_type="solid", start_color='7FFF00', end_color='FF000000'),

    "blue": PatternFill(fill_type="solid", start_color='1E90FF', end_color='FF000000'),

    "yellow": PatternFill(fill_type="solid", start_color='FFD700', end_color='FF000000'),

    "light_blue": PatternFill(fill_type='solid', fgColor='C9D9EE'),
    "dark_blue": PatternFill(fill_type='solid', fgColor='5981B8')

    # 这里添加其他背景颜色

}


# 字体
# 字体可从这里添加
"""
bold  加粗
italic  斜体
strike  删除线
"""
font_styles = {

    #""" 标题字体 """
    "title":  Font(u'Calibri', size=26, bold=True, italic=False, strike=False, color='000000'),


    #"""  内容字体 """
    # 黑色  宋体
    "black":  Font(u'宋体', size=11, bold=False, italic=False, strike=False, color='000000'),

    # 黑色  宋体 加粗
    "black_bold":  Font(u'宋体', size=11, bold=True, italic=False, strike=False, color='000000'),

    # 白色  宋体
    "white": Font(u'宋体', size=11, bold=False, italic=False, strike=False, color='FFFFFF'),
    # 白色  宋体  加粗
    "white_bold": Font(u'宋体', size=11, bold=True, italic=False, strike=False, color='FFFFFF'),

    # 这里添加其他样式字体


}


# 单元格对齐方式
"""
horizontal   水平方向
vertical    垂直方向
"""
cell_alignment = {
    "center": Alignment(
                horizontal='center', vertical='center', text_rotation=0, wrap_text=False, shrink_to_fit=False,
                indent=0
    ),


}


# 单元格 边界 样式
cell_border = {
    "right":  Border(
                left=Side(border_style='thin', color='000000'),
                right=Side(border_style='medium', color='000000'),
                top=Side(border_style='thin', color='000000'),
                bottom=Side(border_style='thin', color='000000')
    )

    # 这里添加其他边界样式
}

# border = Border(left=Side(border_style=None,
#                            color='FF000000'),
#                 right=Side(border_style=None,
#                            color='FF000000'),
#                 top=Side(border_style=None,
#                           color='FF000000'),
#                  bottom=Side(border_style=None,
#                              color='FF000000'),
#                  diagonal=Side(border_style=None,
#                                color='FF000000'),
#                  diagonal_direction=0,
#                  outline=Side(border_style=None,
#                              color='FF000000'),
#                 vertical=Side(border_style=None,
#                               color='FF000000'),
#                 horizontal=Side(border_style=None,
#                                 color='FF000000')
#                )

number_format = 'General'
protection = Protection(locked=True,
                        hidden=False)
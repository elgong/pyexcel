"""
    openpyxl 的封装，有很多函数尽可能的使用原始的名字
"""

import os
import time
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Series, Reference, PieChart3D
from openpyxl.chart.label import DataLabelList

from .Styles import Styles
# from .Config import *

class ExcelUtil(object):

    wb = Workbook()
    # 工作簿的名字
    wb_name = None
    style = Styles()

    def __init__(self, work_book_name=None):
        """
        :param work_book_name:  excel 名
        :param sheet_list:  sheet 名的列表
        """
        if work_book_name is not None:
            self.wb_name = work_book_name

    # 添加表
    def add_sheet(self, sheet_list=[]):
        # 类型检查 -保证传入列表
        if not isinstance(sheet_list, list):
            raise Exception("waring:  sheet_list need  list[] ")
        for sheetName in sheet_list:
            self.wb.create_sheet(sheetName)
        # 删除默认生成的 Sheet
        self.wb.remove(self.wb["Sheet"])

        print("创建sheet列表： ", end="  ")
        print(self.wb.sheetnames)

    # cell 常规设置（表，坐标，值，背景色，字体，边框，对齐方式）
    def set_cell(
                    self,
                    sheet,
                    row,
                    col,
                    value=None,
                    cell_fill=None,
                    cell_border=None,
                    cell_alignment="center",
                    font_type=u'Calibri',
                    font_size=12,
                    font_bold=False,
                    font_italic=False,
                    font_strike=False,
                    font_color="black"
                ):

        cell = self.wb[sheet].cell(row=row, column=col)

        # 设置内容
        if value is not None:
            cell.value = value

        # 填充 背景颜色
        if cell_fill is not None:
            cell.fill = self.style.set_cell_background(cell_fill)

        # 设置字体
        if font_type is not None:
            cell.font = self.style.set_font(font=font_type, size=font_size, bold=font_bold, italic=font_italic, strike=font_strike, color=font_color)

        # 对齐方式
        if cell_alignment is not None:
            cell.alignment = self.style.set_cell_alignment(cell_alignment)

        # # 设置单元格 右边框样式
        # if cell_border == "right":
        #     cell.border = cell_border[border]

        return cell

    def set_title(self):
        pass

    # 调整列宽
    def set_col_weight(self, sheet, col, width=15):
        sheet = self.wb[sheet]
        sheet.column_dimensions[get_column_letter(col)].width = width

    # 调整行高
    def set_row_height(self, sheet, row, height=15):
        sheet = self.wb[sheet]
        sheet.row_dimensions[row].height = height

    # 合并单元格
    def merge_cells(self, sheet, start_row, start_column, end_row, end_column):
        sheet = self.wb[sheet]
        sheet.merge_cells(start_row=start_row, start_column=start_column, end_row=end_row, end_column=end_column)

    # 柱状图
    def draw_bar2D(
                    self,
                    sheet_name,
                    data_position=[],
                    label_position=[],
                    display_position="A10",
                    title="Bar Chart",
                    x_title="display_x",
                    y_title="display_y",
                    is_display_legend=False
                ):
        """
            数据必须为列。。。不知道为啥。。
        """
        ws = self.wb[sheet_name]

        # 数据所在列的坐标范围， 不包含label
        DATA_COL_MIN = data_position[0]
        DATA_COL_MAX = data_position[1]
        DATA_ROW_MIN = data_position[2]
        DATA_ROW_MAX = data_position[3]

        # label 所在范围
        LABEL_COL_MIN = label_position[0]
        LABEL_COL_MAX = label_position[1]
        LABEL_ROW_MIN = label_position[2]
        LABEL_ROW_MAX = label_position[3]

        # 创建 chart 对象
        chart1 = BarChart()

        # 竖直的柱状图"col"
        chart1.type = "col"
        chart1.style = 10
        chart1.shape = 4
        chart1.title = title
        chart1.y_axis.title = y_title
        chart1.x_axis.title = x_title

        data = Reference(ws, min_col=DATA_COL_MIN, max_col=DATA_COL_MAX, min_row=DATA_ROW_MIN, max_row=DATA_ROW_MAX)
        cats = Reference(ws, min_col=LABEL_COL_MIN, max_col=LABEL_COL_MAX, min_row=LABEL_ROW_MIN, max_row=LABEL_ROW_MAX)

        chart1.add_data(data, titles_from_data=True)
        chart1.set_categories(cats)

        # label , 柱状图上的数字
        chart1.dLbls = DataLabelList()
        chart1.dLbls.showVal = True

        # 是否显示图例图例
        if is_display_legend == False:
            chart1.legend = None

        # 显示位置
        ws.add_chart(chart1, display_position)

    # 饼状图
    def draw_pie3D(
                    self,
                    sheet_name,
                    data_position=[],
                    label_position=[],
                    display_position="A10",
                    title="Bar Chart",
                    x_title="display_x",
                    y_title="display_y",
                    is_display_legend=False
            ):
        """
            数据必须为列。。。不知道为啥。。
        """
        ws = self.wb[sheet_name]

        # 数据所在列的坐标范围， 不包含label
        DATA_COL_MIN = data_position[0]
        DATA_COL_MAX = data_position[1]
        DATA_ROW_MIN = data_position[2]
        DATA_ROW_MAX = data_position[3]

        # label 所在范围
        LABEL_COL_MIN = label_position[0]
        LABEL_COL_MAX = label_position[1]
        LABEL_ROW_MIN = label_position[2]
        LABEL_ROW_MAX = label_position[3]

        chart1 = PieChart3D()

        # 竖直的柱状图"col"
        chart1.type = "col"
        chart1.style = 10
        chart1.shape = 4
        chart1.title = title

        data = Reference(ws, min_col=DATA_COL_MIN, max_col=DATA_COL_MAX, min_row=DATA_ROW_MIN, max_row=DATA_ROW_MAX)
        cats = Reference(ws, min_col=LABEL_COL_MIN, max_col=LABEL_COL_MAX, min_row=LABEL_ROW_MIN, max_row=LABEL_ROW_MAX)

        chart1.add_data(data, titles_from_data=True)
        chart1.set_categories(cats)

        # label , 柱状图上的数字
        chart1.dLbls = DataLabelList()
        chart1.dLbls.showVal = True

        # 显示位置
        ws.add_chart(chart1, display_position)

    # 自动保存文件
    def save(self):
        try:
            if os.path.exists(self.wb_name):
                os.remove(self.wb_name)
            self.wb.save(self.wb_name)
            print("save success")
        except PermissionError as e:
            print("Permission denied!, Wait close the old file")
            print("You have 5s to close")
            time.sleep(5)
            self.wb.save(self.wb_name)
            print("save success")















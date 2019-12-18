"""

"""

import os
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from pyexcel.Config import *


class ExcelUtil(object):

    obj = None
    wb = Workbook()
    # 工作簿的名字
    wb_name = None

    def __init__(self, work_book_name=None, sheet_list=["sheet1"]):
        """
        :param work_book_name:  excel 名
        :param sheet_list:  sheet 名的列表
        """
        if work_book_name is not None:
            self.wb_name = work_book_name
            # 创建 sheet
            for sheetName in sheet_list:
                self.wb.create_sheet(sheetName)

            # 删除默认生成的 Sheet
            self.wb.remove(self.wb["Sheet"])

            print("创建sheet列表： ", end="  ")
            print(self.wb.sheetnames)

            self.wb.save(self.wb_name)

    # 返回 wb对象
    def get_wb(self):
        """
          return:
        """
        return self.wb

    # cell 设置
    def set_cell(self, sheet="sheet1", row=1, col=1, value="", background=None, font_color=None, alignment=None):
        cell = self.wb[sheet].cell(row=row, column=col, value=value)

        # 设置 背景颜色
        if background is not None:
            cell.fill = cell_background[background]

        # 设置字体
        if font_color is not None:
            cell.font = font_styles[font_color]

        # 单元格对齐方式
        if alignment is not None:
            cell.alignment = cell_alignment[alignment]

    # 保存文件
    def save_wb(self):
        self.wb.save(self.wb_name)


if __name__ == "__main__":

    excelUtil = ExcelUtil("../output/123.xlsx")
    excelUtil.set_cell(sheet="sheet1", row=1, col=2, value="123", background="red")
    excelUtil.save_wb()










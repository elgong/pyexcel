"""
    log：
    2020/1/8
    重构了格式控制代码。
    通过Styles 类提供所有的格式风格。
"""
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font


class Styles(object):

    """ 预设参数区 """
    # 背景颜色配置
    # fill_type = "solid"  # 纯色填充方式
    # start_color = 'FF0000FF'  # 前景色
    # end_color = 'FF000000'  # 背景色
    cell_background_parameter = {
        None: PatternFill(fill_type="none", start_color='FFFFFF', end_color='FF000000'),

        "red": PatternFill(fill_type="solid", start_color='FF4500', end_color='FF000000'),

        "green": PatternFill(fill_type="solid", start_color='7FFF00', end_color='FF000000'),

        "blue": PatternFill(fill_type="solid", start_color='1E90FF', end_color='FF000000'),

        "yellow": PatternFill(fill_type="solid", start_color='FFD700', end_color='FF000000'),

        "light_blue": PatternFill(fill_type='solid', fgColor='C9D9EE'),

        "dark_blue": PatternFill(fill_type='solid', fgColor='5981B8')
        # 在这里添加其他背景颜色，颜色表可查 http://www.365jz.com/article/24452
    }

    # 单元格对齐方式
    # 居中，自动换行
    cell_alignment = {
        "center": Alignment(
            horizontal='center', vertical='center', text_rotation=0, wrap_text=True, shrink_to_fit=False,
            indent=0
        ),
    }

    def __init__(self):
        pass

    # 单元格背景颜色
    def set_cell_background(self, background_color=None):
        try:
            return self.cell_background_parameter[background_color]
        except:
            raise Exception("Background_color Error!")

    # 字体风格
    def set_font(self, font=u'Calibri', size=12, bold=False, italic=False, strike=False, color="black"):
        """
        功能: 字体风格设置, 字体，大小，加粗，斜体， 下划线，颜色等
            参数名: [可选参数,]
            _font: [u'宋体', u'Calibri']
            _size:    字体大小
            _bold:    是否加粗
            _italic:  斜体
            _strike: 下划线
            _color:  字体颜色
        """
        if color == "black":  # 默认白色
            color_para = "000000"
        elif color == "white":
            color_para ="FFFFFF"
        else:
            raise Exception("Font color Error !")

        return Font(font, size=size, bold=bold, italic=italic, strike=strike, color=color_para)

    # 单元格对齐
    def set_cell_alignment(self, alignment="center"):
        try:
            return self.cell_alignment[alignment]
        except:
            raise Exception("Alignment Error ！")




